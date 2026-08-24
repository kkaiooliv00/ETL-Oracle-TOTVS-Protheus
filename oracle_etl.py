from __future__ import annotations
import argparse
import json
import logging
import os
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import oracledb
import pandas as pd
from pandas.api.types import is_string_dtype, is_object_dtype
import psycopg
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import Connection, Engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.pool import QueuePool

load_dotenv()

# ── Constantes Oracle ─────────────────────────────────────────────────────────
ORA_HOST_ENV = "ORA_HOST"
ORA_PORT_ENV = "ORA_PORT"
ORA_SERVICE_ENV = "ORA_SERVICE"
ORA_USER_ENV = "ORA_USER"
ORA_PASSWORD_ENV = "ORA_PASSWORD"

# ── Constantes PostgreSQL ─────────────────────────────────────────────────────
DATABASE_URL_ENV = "DATABASE_URL"

# ── Config ────────────────────────────────────────────────────────────────────
JOB_CONFIG_PATH = Path(os.getenv("ETL_JOBS_FILE", Path(__file__).parent / "oracle_jobs.yml"))

# Registros por batch do cursor Oracle.
FETCH_SIZE = 10_000

# Registros acumulados antes de despejar no staging via COPY.
STAGING_FLUSH_RECORDS = 10_000

BUSINESS_TIMEZONE = ZoneInfo("America/Fortaleza")
TARGET_SCHEMA = "tables"
BUSINESS_KEY = "super_chave"

# Filtro de exclusao logica do TOTVS.
TOTVS_DELETED_FILTER = "D_E_L_E_T_ <> '*'"

# Tentativas de retry por job.
JOB_MAX_ATTEMPTS = 3

# Limite de linhas por aba imposto pelo formato XLSX.
EXCEL_MAX_ROWS_PER_SHEET = 1_048_576
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("oracle_etl")


# ── Modelos ───────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class EtlJob:
    oracle_table: str
    target_table: str
    query: str
    date_column: str | None = None
    business_key: str = BUSINESS_KEY
    business_key_columns: tuple[str, ...] = ()

    @property
    def staging_table(self) -> str:
        return f"{self.target_table}_staging"

    @property
    def dedup_staging_table(self) -> str:
        return f"{self.target_table}_staging_dedup"


# ── Utilitarios ───────────────────────────────────────────────────────────────

def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Variavel de ambiente obrigatoria nao definida: {name}")
    return value


def require_postgres_database_url() -> str:
    database_url = require_env(DATABASE_URL_ENV)
    scheme = urlparse(database_url).scheme
    if scheme not in ("postgresql", "postgresql+psycopg", "postgres"):
        raise RuntimeError(
            "DATABASE_URL invalida. Use uma URL PostgreSQL como "
            "postgresql+psycopg://usuario:senha@host:6543/postgres."
        )
    return database_url


def _normalize_dsn(database_url: str) -> str:
    return (
        database_url
        .replace("postgresql+psycopg://", "postgresql://")
        .replace("postgresql+psycopg2://", "postgresql://")
    )


def _sqlalchemy_database_url(database_url: str) -> str:
    if database_url.startswith("postgresql://"):
        return database_url.replace("postgresql://", "postgresql+psycopg://", 1)
    if database_url.startswith("postgres://"):
        return database_url.replace("postgres://", "postgresql+psycopg://", 1)
    return database_url


def preflight_database_connection(engine: Engine) -> None:
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))


def quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def qualified_table(schema: str, table: str) -> str:
    return f"{quote_identifier(schema)}.{quote_identifier(table)}"


# ── Carregamento de jobs ──────────────────────────────────────────────────────

def load_jobs() -> tuple[str, list[EtlJob]]:
    """Carrega a configuracao de jobs do YAML.

    Retorna uma tupla (oracle_schema, lista_de_jobs).
    """
    if not JOB_CONFIG_PATH.exists():
        raise RuntimeError(f"Arquivo de jobs nao encontrado: {JOB_CONFIG_PATH}")

    with JOB_CONFIG_PATH.open(encoding="utf-8") as f:
        config = yaml.safe_load(f) or {}

    oracle_schema = str(config.get("oracle_schema", "")).strip()
    if not oracle_schema:
        raise RuntimeError("oracle_jobs.yml deve conter oracle_schema.")

    raw_jobs = config.get("jobs")
    if not isinstance(raw_jobs, list) or not raw_jobs:
        raise RuntimeError("oracle_jobs.yml deve conter uma lista nao vazia em jobs.")

    jobs: list[EtlJob] = []
    for raw in raw_jobs:
        if not isinstance(raw, dict):
            raise RuntimeError("Cada job deve conter oracle_table e target_table.")
        try:
            oracle_table = str(raw["oracle_table"]).strip()
            target_table = str(raw["target_table"]).strip()
            query = raw.get("query")
            date_column = raw.get("date_column")
            business_key = str(raw.get("business_key", BUSINESS_KEY)).strip()
            raw_bk_cols = raw.get("business_key_columns") or []
        except (KeyError, TypeError, ValueError) as exc:
            raise RuntimeError(f"Job invalido em {JOB_CONFIG_PATH}: {raw!r}") from exc

        if not oracle_table:
            raise RuntimeError("oracle_table nao pode ser vazio.")
        if not target_table:
            raise RuntimeError("target_table nao pode ser vazio.")
        if not business_key:
            raise RuntimeError(f"business_key vazio em {target_table}.")
        if query is None or not str(query).strip():
            raise RuntimeError(
                f"Job '{target_table}' nao possui campo 'query' no YAML. "
                f"Toda tabela DEVE ter uma query personalizada. "
                f"SELECT * nao e permitido."
            )
        query = str(query).strip()
        if date_column is not None:
            date_column = str(date_column).strip()
        business_key_columns = tuple(
            str(c).strip() for c in raw_bk_cols if str(c).strip()
        )
        jobs.append(EtlJob(
            oracle_table=oracle_table,
            target_table=target_table,
            query=query,
            date_column=date_column,
            business_key=business_key,
            business_key_columns=business_key_columns,
        ))
    return oracle_schema, jobs


# ── Conexao Oracle ────────────────────────────────────────────────────────────

def create_oracle_connection() -> oracledb.Connection:
    """Cria uma conexao com o Oracle TOTVS Protheus.

    Usa o driver oracledb em modo thin (Python puro — nao precisa de
    Oracle Instant Client instalado).
    """
    host = require_env(ORA_HOST_ENV)
    port = require_env(ORA_PORT_ENV)
    service = require_env(ORA_SERVICE_ENV)
    user = require_env(ORA_USER_ENV)
    password = require_env(ORA_PASSWORD_ENV)

    dsn = oracledb.makedsn(host, int(port), service_name=service)
    conn = oracledb.connect(user=user, password=password, dsn=dsn)
    logger.info("Conexao Oracle estabelecida: %s@%s:%s/%s", user, host, port, service)
    return conn


def preflight_oracle_connection() -> None:
    """Valida a conexao com o Oracle antes de iniciar o pipeline."""
    conn = create_oracle_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM DUAL")
            cur.fetchone()
        logger.info("Preflight Oracle: conexao OK.")
    finally:
        conn.close()


# ── Extracao Oracle ───────────────────────────────────────────────────────────

def format_start_date(lookback_days: int) -> str:
    """Retorna a data de inicio no formato YYYYMMDD do TOTVS."""
    current_date = datetime.now(BUSINESS_TIMEZONE).date()
    return (current_date - timedelta(days=lookback_days)).strftime("%Y%m%d")


def _build_query(
    oracle_schema: str,
    job: EtlJob,
    lookback_days: int | None,
) -> tuple[str, dict[str, Any]]:
    """Prepara a query customizada para execucao.

    Substitui o placeholder {schema} pelo oracle_schema real.
    Toda query DEVE ser definida no YAML — SELECT * nao e permitido.

    Retorna (sql_string, bind_params).
    """
    params: dict[str, Any] = {}

    if not job.query:
        raise RuntimeError(
            f"Job '{job.target_table}' nao possui query personalizada. "
            f"SELECT * nao e permitido. Defina o campo 'query' no YAML."
        )

    sql = job.query.replace("{schema}", oracle_schema)
    logger.info("%s | Query: %s", job.target_table, sql)
    return sql, params


def iter_oracle_batches(
    ora_conn: oracledb.Connection,
    oracle_schema: str,
    job: EtlJob,
    lookback_days: int | None,
) -> Iterator[tuple[int, list[dict[str, Any]]]]:
    """Extrai registros do Oracle em batches via cursor.fetchmany().

    Retorna Iterator de (batch_number, lista_de_dicts) — mesmo contrato
    do iter_api_pages() do ETL antigo para manter compatibilidade com
    o run_job().
    """
    sql, params = _build_query(oracle_schema, job, lookback_days)
    total_extracted = 0

    with ora_conn.cursor() as cursor:
        cursor.prefetchrows = FETCH_SIZE + 1
        cursor.arraysize = FETCH_SIZE
        cursor.execute(sql, params)

        # Obtem nomes de colunas do cursor para converter rows em dicts.
        col_names = [desc[0] for desc in cursor.description]
        # Converte nomes para lowercase para manter consistencia com o
        # ETL existente (Oracle retorna uppercase por padrao).
        col_names_lower = [c.lower() for c in col_names]

        batch_num = 0
        while True:
            rows = cursor.fetchmany(FETCH_SIZE)
            batch_num += 1

            if not rows:
                logger.info(
                    "%s | Extracao Oracle concluida: %s registros totais em %s batches.",
                    job.target_table, total_extracted, batch_num - 1,
                )
                # Yield final vazio para sinalizar fim (compativel com run_job)
                yield batch_num, []
                break

            records = [dict(zip(col_names_lower, row)) for row in rows]
            total_extracted += len(records)

            logger.info(
                "%s | Batch %s: %s registros (total acumulado: %s).",
                job.target_table, batch_num, len(records), total_extracted,
            )

            yield batch_num, records


# ── Transformacao ─────────────────────────────────────────────────────────────

def transform_records(job: EtlJob, records: list[dict[str, Any]]) -> pd.DataFrame:
    """Normaliza os registros e garante a presenca da business key quando possivel.

    Se a business_key nao existir nos dados E nao houver business_key_columns
    configuradas, o DataFrame e retornado sem a coluna de chave. Nesse caso,
    finalize_load realizara um INSERT total (sem upsert) na tabela destino.

    A deduplicacao e feita em uma unica etapa no banco de dados via
    create_dedup_staging() (ROW_NUMBER + ctid DESC), antes do upsert final.
    """
    df = pd.json_normalize(records, sep="_")
    if df.empty:
        return df

    # Limpeza de campos textuais/código vindos do Protheus (remove CHR(160) e trailing spaces)
    # Isso evita problemas com campos fixos preenchidos com espaço, garantindo a integridade dos joins e filtros.
    for col in df.columns:
        if is_object_dtype(df[col]) or is_string_dtype(df[col]):
            df[col] = df[col].apply(
                lambda v: v.replace("\xa0", " ").strip() if isinstance(v, str) else v
            )

    if job.business_key not in df.columns and job.business_key_columns:
        missing = [c for c in job.business_key_columns if c not in df.columns]
        if missing:
            raise KeyError(
                f"Colunas para chave composta ausentes em {job.target_table}: "
                f"{', '.join(missing)}"
            )
        df[job.business_key] = (
            df.loc[:, list(job.business_key_columns)]
            .fillna("")
            .astype(str)
            .agg("|".join, axis=1)
        )

    if job.business_key not in df.columns:
        logger.warning(
            "%s | business_key '%s' ausente nos dados e sem business_key_columns "
            "configuradas. Carga sera realizada como INSERT total (sem upsert). "
            "Colunas recebidas: %s",
            job.target_table, job.business_key,
            ", ".join(str(c) for c in df.columns),
        )

    return df


# ── Escrita no staging via COPY (psycopg3) ────────────────────────────────────

def _prepare_dataframe_for_copy(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        if is_object_dtype(df[col]) or is_string_dtype(df[col]):
            # Remove NUL bytes (\x00) que o PostgreSQL nao aceita em texto.
            df[col] = df[col].apply(
                lambda v: v.replace("\x00", "") if isinstance(v, str) else v
            )
            sample = df[col].dropna().head(10)
            if any(isinstance(v, (dict, list)) for v in sample):
                df[col] = df[col].apply(
                    lambda v: json.dumps(v, ensure_ascii=False)
                    if isinstance(v, (dict, list)) else v
                )
    return df


def _copy_dataframe_to_staging(
    dsn: str,
    job: EtlJob,
    df: pd.DataFrame,
    if_exists: str,
) -> None:
    if df.empty:
        return

    df = _prepare_dataframe_for_copy(df)
    staging_fqn = qualified_table(TARGET_SCHEMA, job.staging_table)
    columns_sql = ", ".join(quote_identifier(c) for c in df.columns)

    with psycopg.connect(dsn) as conn:
        if if_exists == "replace":
            col_defs = []
            for col in df.columns:
                dtype = str(df[col].dtype)
                if "int" in dtype or "float" in dtype:
                    pg_type = "NUMERIC"
                elif "bool" in dtype:
                    pg_type = "BOOLEAN"
                elif "datetime" in dtype:
                    pg_type = "TIMESTAMPTZ"
                else:
                    pg_type = "TEXT"
                col_defs.append(f"{quote_identifier(col)} {pg_type}")

            conn.execute(f"DROP TABLE IF EXISTS {staging_fqn}")
            conn.execute(f"CREATE TABLE {staging_fqn} ({', '.join(col_defs)})")

        copy_sql = (
            f"COPY {staging_fqn} ({columns_sql}) "
            f"FROM STDIN WITH (FORMAT text, DELIMITER E'\\t', NULL '\\N')"
        )

        with conn.cursor() as cur:
            with cur.copy(copy_sql) as copy:
                for row in df.itertuples(index=False, name=None):
                    formatted: list[str] = []
                    for val in row:
                        if val is None or val is pd.NaT or (isinstance(val, float) and pd.isna(val)):
                            formatted.append(None)
                        else:
                            formatted.append(
                                str(val)
                                .replace("\x00", "")
                                .replace("\\", "\\\\")
                                .replace("\t", "\\t")
                                .replace("\n", "\\n")
                                .replace("\r", "\\r")
                            )
                    copy.write_row(formatted)

        conn.commit()

    logger.info(
        "%s | COPY concluido: %s registros -> %s.",
        job.target_table, len(df), job.staging_table,
    )


# ── DDL e Upsert ──────────────────────────────────────────────────────────────

def create_schema(connection: Connection) -> None:
    connection.execute(
        text(f"CREATE SCHEMA IF NOT EXISTS {quote_identifier(TARGET_SCHEMA)}")
    )


def drop_staging_tables(connection: Connection, job: EtlJob) -> None:
    connection.execute(
        text(f"DROP TABLE IF EXISTS {qualified_table(TARGET_SCHEMA, job.dedup_staging_table)}")
    )
    connection.execute(
        text(f"DROP TABLE IF EXISTS {qualified_table(TARGET_SCHEMA, job.staging_table)}")
    )


def create_target_from_staging(connection: Connection, job: EtlJob) -> None:
    target = qualified_table(TARGET_SCHEMA, job.target_table)
    staging = qualified_table(TARGET_SCHEMA, job.staging_table)
    connection.execute(
        text(f"CREATE TABLE IF NOT EXISTS {target} AS TABLE {staging} WITH NO DATA")
    )


def add_missing_target_columns(
    connection: Connection, engine: Engine, job: EtlJob
) -> None:
    inspector = inspect(connection)
    target_columns = {
        col["name"]
        for col in inspector.get_columns(job.target_table, schema=TARGET_SCHEMA)
    }
    for col in inspector.get_columns(job.staging_table, schema=TARGET_SCHEMA):
        col_name = col["name"]
        if col_name in target_columns:
            continue
        col_type = col["type"].compile(dialect=engine.dialect)
        connection.execute(
            text(
                f"ALTER TABLE {qualified_table(TARGET_SCHEMA, job.target_table)} "
                f"ADD COLUMN IF NOT EXISTS {quote_identifier(col_name)} {col_type}"
            )
        )


def deduplicate_target_table(connection: Connection, job: EtlJob) -> None:
    """Remove duplicatas da tabela alvo antes de criar a UNIQUE constraint.

    Necessario quando a tabela ja existia sem constraint e acumulou
    duplicatas em execucoes anteriores. Mantem o registro de menor ctid
    (o mais antigo fisicamente) para cada business_key.
    Se a tabela ainda nao existir, nao faz nada.
    """
    target = qualified_table(TARGET_SCHEMA, job.target_table)
    key = quote_identifier(job.business_key)

    # Verifica se a tabela existe antes de tentar limpar
    table_exists = connection.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1
                FROM pg_class t
                JOIN pg_namespace n ON n.oid = t.relnamespace
                WHERE t.relname = :tn AND n.nspname = :sn
            )
            """
        ),
        {"tn": job.target_table, "sn": TARGET_SCHEMA},
    ).scalar_one()

    if not table_exists:
        return

    result = connection.execute(
        text(
            f"""
            DELETE FROM {target} t1
            USING {target} t2
            WHERE t1.{key} = t2.{key}
              AND t1.ctid > t2.ctid
            """
        )
    )
    if result.rowcount > 0:
        logger.warning(
            "%s | %s registros duplicados removidos da tabela alvo antes de criar constraint.",
            job.target_table, result.rowcount,
        )


def ensure_unique_constraint(connection: Connection, job: EtlJob) -> None:
    target = qualified_table(TARGET_SCHEMA, job.target_table)
    constraint_name = f"{job.target_table}_{job.business_key}_uk"
    key = quote_identifier(job.business_key)

    exists = connection.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1
                FROM pg_constraint c
                JOIN pg_class t ON t.oid = c.conrelid
                JOIN pg_namespace n ON n.oid = t.relnamespace
                WHERE c.conname = :cn AND t.relname = :tn AND n.nspname = :sn
            )
            """
        ),
        {"cn": constraint_name, "tn": job.target_table, "sn": TARGET_SCHEMA},
    ).scalar_one()

    if not exists:
        connection.execute(
            text(
                f"ALTER TABLE {target} ADD CONSTRAINT "
                f"{quote_identifier(constraint_name)} UNIQUE ({key})"
            )
        )


def create_dedup_staging(
    connection: Connection, job: EtlJob, columns: list[str]
) -> None:
    quoted_cols = ", ".join(quote_identifier(c) for c in columns)
    source = qualified_table(TARGET_SCHEMA, job.staging_table)
    dedup = qualified_table(TARGET_SCHEMA, job.dedup_staging_table)
    key = quote_identifier(job.business_key)

    connection.execute(text(f"DROP TABLE IF EXISTS {dedup}"))
    connection.execute(
        text(
            f"""
            CREATE TABLE {dedup} AS
            SELECT {quoted_cols}
            FROM (
                SELECT {quoted_cols},
                       ROW_NUMBER() OVER (
                           PARTITION BY {key} ORDER BY ctid DESC
                       ) AS __rn
                FROM {source}
            ) s
            WHERE __rn = 1
            """
        )
    )

    count_res = connection.execute(text(f"SELECT count(*) FROM {dedup}"))
    count = count_res.scalar_one()
    logger.info("%s | Staging deduplicado criado com %s registros.", job.target_table, count)


def upsert_from_staging(
    connection: Connection, job: EtlJob, columns: list[str]
) -> None:
    quoted_cols = ", ".join(quote_identifier(c) for c in columns)
    update_cols = [c for c in columns if c != job.business_key]
    key = quote_identifier(job.business_key)
    dedup = qualified_table(TARGET_SCHEMA, job.dedup_staging_table)
    target = qualified_table(TARGET_SCHEMA, job.target_table)

    if update_cols:
        assignments = ", ".join(
            f"{quote_identifier(c)} = EXCLUDED.{quote_identifier(c)}"
            for c in update_cols
        )
        conflict_action = f"DO UPDATE SET {assignments}"
    else:
        conflict_action = "DO NOTHING"

    result = connection.execute(
        text(
            f"""
            INSERT INTO {target} ({quoted_cols})
            SELECT {quoted_cols} FROM {dedup}
            ON CONFLICT ({key}) {conflict_action}
            """
        )
    )
    logger.info(
        "%s | UPSERT concluido: %s linhas afetadas.",
        job.target_table, result.rowcount,
    )


def insert_all_from_staging(
    connection: Connection, job: EtlJob, columns: list[str]
) -> None:
    """Realiza INSERT total do staging para a tabela destino (sem upsert).

    Utilizado quando a tabela nao possui super_chave. A tabela destino e
    truncada antes do INSERT para evitar duplicatas entre execucoes.
    """
    quoted_cols = ", ".join(quote_identifier(c) for c in columns)
    staging = qualified_table(TARGET_SCHEMA, job.staging_table)
    target = qualified_table(TARGET_SCHEMA, job.target_table)

    connection.execute(text(f"TRUNCATE TABLE {target}"))
    result = connection.execute(
        text(
            f"INSERT INTO {target} ({quoted_cols}) "
            f"SELECT {quoted_cols} FROM {staging}"
        )
    )
    logger.info(
        "%s | INSERT total concluido (sem super_chave): %s linhas inseridas.",
        job.target_table, result.rowcount,
    )


def _staging_has_business_key(connection: Connection, job: EtlJob) -> bool:
    """Verifica se a coluna business_key existe na tabela de staging."""
    inspector = inspect(connection)
    staging_columns = {
        col["name"]
        for col in inspector.get_columns(job.staging_table, schema=TARGET_SCHEMA)
    }
    return job.business_key in staging_columns


def delete_from_staging(connection: Connection, job: EtlJob) -> None:
    key = quote_identifier(job.business_key)
    staging = qualified_table(TARGET_SCHEMA, job.staging_table)
    target = qualified_table(TARGET_SCHEMA, job.target_table)
    
    table_exists = connection.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1
                FROM pg_class t
                JOIN pg_namespace n ON n.oid = t.relnamespace
                WHERE t.relname = :tn AND n.nspname = :sn
            )
            """
        ),
        {"tn": job.target_table, "sn": TARGET_SCHEMA},
    ).scalar_one()

    if not table_exists:
        logger.info("%s | Tabela alvo nao existe no destino. Nada a excluir.", job.target_table)
        return
        
    result = connection.execute(
        text(
            f"DELETE FROM {target} WHERE {key} IN (SELECT {key} FROM {staging})"
        )
    )
    logger.info(
        "%s | DELETE concluido: %s linhas excluidas fisicamente do Supabase.",
        job.target_table, result.rowcount,
    )


def finalize_load(engine: Engine, job: EtlJob, columns: list[str], is_delete_mode: bool = False) -> None:
    logger.info("%s | Iniciando finalize_load.", job.target_table)
    try:
        with engine.begin() as conn:
            if is_delete_mode:
                t0 = time.perf_counter()
                delete_from_staging(conn, job)
                logger.info("%s | delete_from_staging: %.1fs.", job.target_table, time.perf_counter() - t0)
            else:
                t0 = time.perf_counter()
                create_target_from_staging(conn, job)
                logger.info("%s | create_target_from_staging: %.1fs.", job.target_table, time.perf_counter() - t0)

                t0 = time.perf_counter()
                add_missing_target_columns(conn, engine, job)
                logger.info("%s | add_missing_target_columns: %.1fs.", job.target_table, time.perf_counter() - t0)

                has_key = _staging_has_business_key(conn, job)

                if not has_key:
                    # Tabela sem super_chave: realiza INSERT total (TRUNCATE + INSERT)
                    logger.info(
                        "%s | super_chave '%s' ausente na tabela de staging. "
                        "Executando INSERT total (TRUNCATE + INSERT).",
                        job.target_table, job.business_key,
                    )
                    t0 = time.perf_counter()
                    insert_all_from_staging(conn, job, columns)
                    logger.info("%s | insert_all_from_staging: %.1fs.", job.target_table, time.perf_counter() - t0)
                else:
                    # Tabela com super_chave: fluxo normal de upsert
                    t0 = time.perf_counter()
                    deduplicate_target_table(conn, job)
                    logger.info("%s | deduplicate_target_table: %.1fs.", job.target_table, time.perf_counter() - t0)

                    t0 = time.perf_counter()
                    ensure_unique_constraint(conn, job)
                    logger.info("%s | ensure_unique_constraint: %.1fs.", job.target_table, time.perf_counter() - t0)

                    t0 = time.perf_counter()
                    create_dedup_staging(conn, job, columns)
                    logger.info("%s | create_dedup_staging: %.1fs.", job.target_table, time.perf_counter() - t0)

                    t0 = time.perf_counter()
                    upsert_from_staging(conn, job, columns)
                    logger.info("%s | upsert_from_staging: %.1fs.", job.target_table, time.perf_counter() - t0)

            t0 = time.perf_counter()
            drop_staging_tables(conn, job)
            logger.info("%s | drop_staging_tables: %.1fs.", job.target_table, time.perf_counter() - t0)

    except Exception:
        logger.exception(
            "%s | Falha em finalize_load. Tabelas de staging preservadas para inspecao: %s, %s.",
            job.target_table, job.staging_table, job.dedup_staging_table,
        )
        raise


# ── Orquestrador por job ──────────────────────────────────────────────────────

def run_job(
    engine: Engine,
    ora_conn: oracledb.Connection,
    oracle_schema: str,
    job: EtlJob,
    lookback_days: int | None,
    is_delete_mode: bool = False,
) -> None:
    started_at = time.perf_counter()
    logger.info(
        "Iniciando job oracle_table=%s target_table=%s.",
        job.oracle_table, job.target_table,
    )

    dsn = _normalize_dsn(require_env(DATABASE_URL_ENV))
    extracted_records = 0
    staged_records = 0
    columns: list[str] | None = None
    staging_mode = "replace"
    buffer_frames: list[pd.DataFrame] = []
    buffer_records = 0

    with engine.begin() as conn:
        create_schema(conn)
        drop_staging_tables(conn, job)

    try:
        for batch_num, records in iter_oracle_batches(ora_conn, oracle_schema, job, lookback_days):
            extracted_records += len(records)
            if not records:
                continue

            df = transform_records(job, records)
            if df.empty:
                continue

            # Acumula a uniao de todas as colunas vistas ate agora.
            new_cols = [c for c in df.columns if c not in (columns or [])]
            if new_cols:
                columns = list(columns or []) + new_cols

            buffer_frames.append(df)
            buffer_records += len(df)

            if buffer_records >= STAGING_FLUSH_RECORDS:
                merged = pd.concat(buffer_frames, ignore_index=True)
                _copy_dataframe_to_staging(dsn, job, merged, staging_mode)
                staged_records += len(merged)
                staging_mode = "append"
                buffer_frames.clear()
                buffer_records = 0

    except SQLAlchemyError:
        logger.exception("%s | Falha durante a carga no PostgreSQL.", job.target_table)
        raise

    if columns is None:
        logger.info(
            "%s | Nenhum registro retornado; carga dispensada.", job.target_table
        )
        return

    # Flush do buffer residual
    if buffer_frames:
        merged = pd.concat(buffer_frames, ignore_index=True)
        _copy_dataframe_to_staging(dsn, job, merged, staging_mode)
        staged_records += len(merged)

    finalize_load(engine, job, columns, is_delete_mode)

    elapsed = time.perf_counter() - started_at
    records_per_sec = extracted_records / elapsed if elapsed > 0 else 0
    logger.info(
        "%s | Carga concluida em %s.%s | extraidos=%s staging=%s tempo=%.1fs velocidade=%.0f reg/s.",
        job.target_table, TARGET_SCHEMA, job.target_table,
        extracted_records, staged_records, elapsed, records_per_sec
    )


# ── Exportacao XLSX ────────────────────────────────────────────────────────

def _excel_safe_sheet_name(name: str, used_names: set[str], part: int) -> str:
    """Gera um nome de aba compativel com o Excel e sem duplicidade."""
    base = "".join("_" if char in "[]:*?/\\\\" else char for char in name).strip()
    base = base or "dados"
    suffix = "" if part == 1 else f"_{part}"
    candidate = f"{base[:31 - len(suffix)]}{suffix}"
    while candidate.lower() in used_names:
        part += 1
        suffix = f"_{part}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
    used_names.add(candidate.lower())
    return candidate


def _excel_value(value: Any) -> Any:
    """Converte valores do Oracle/Pandas para tipos aceitos pelo XLSX."""
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, str):
        return value.replace("\x00", "")
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def export_jobs_to_excel(
    ora_conn: oracledb.Connection,
    oracle_schema: str,
    jobs: list[EtlJob],
    lookback_days: int | None,
    output_file: Path,
) -> None:
    """Extrai os jobs selecionados para um unico XLSX, sem usar PostgreSQL."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    used_sheet_names: set[str] = set()
    total_records = 0

    try:
        for job in jobs:
            sheet_part = 1
            worksheet = None
            columns: list[str] | None = None
            rows_in_sheet = 0
            job_records = 0

            for _, records in iter_oracle_batches(ora_conn, oracle_schema, job, lookback_days):
                if not records:
                    continue
                df = transform_records(job, records)
                if df.empty:
                    continue

                if columns is None:
                    columns = [str(column) for column in df.columns]
                extra_columns = [column for column in df.columns if column not in columns]
                if extra_columns:
                    raise RuntimeError(
                        f"{job.target_table} retornou colunas diferentes entre batches: "
                        f"{', '.join(str(column) for column in extra_columns)}"
                    )

                df = df.reindex(columns=columns)
                for row in df.itertuples(index=False, name=None):
                    if worksheet is None or rows_in_sheet >= EXCEL_MAX_ROWS_PER_SHEET:
                        sheet_name = _excel_safe_sheet_name(
                            job.target_table, used_sheet_names, sheet_part
                        )
                        sheet_part += 1
                        worksheet = workbook.create_sheet(sheet_name)
                        worksheet.append(columns)
                        rows_in_sheet = 1
                    worksheet.append([_excel_value(value) for value in row])
                    rows_in_sheet += 1
                    job_records += 1
                    total_records += 1

            if columns is None:
                worksheet = workbook.create_sheet(
                    _excel_safe_sheet_name(job.target_table, used_sheet_names, sheet_part)
                )
                worksheet.append(["mensagem"])
                worksheet.append(["Nenhum registro retornado pela consulta."])

            logger.info(
                "%s | Exportacao XLSX concluida: %s registros.",
                job.target_table, job_records,
            )

        workbook.save(output_file)
    except Exception:
        workbook.close()
        raise

    logger.info(
        "Arquivo XLSX gerado com sucesso: %s | %s registros em %s jobs.",
        output_file, total_records, len(jobs),
    )


# ── CLI e entrypoint ──────────────────────────────────────────────────────────

def filter_jobs_by_tables(
    jobs: list[EtlJob], tables: set[str]
) -> list[EtlJob]:
    """Retorna apenas os jobs cujo target_table ou oracle_table esta na lista."""
    if not tables:
        return jobs
    tables_lower = {t.lower() for t in tables}
    selected = [
        j for j in jobs
        if j.target_table.lower() in tables_lower
        or j.oracle_table.lower() in tables_lower
    ]
    if not selected:
        available = ", ".join(j.target_table for j in jobs)
        raise RuntimeError(
            f"Nenhum job encontrado para --tables={sorted(tables)}. "
            f"Tabelas disponiveis: {available}"
        )
    return selected


def exclude_jobs_by_tables(
    jobs: list[EtlJob], tables: set[str]
) -> list[EtlJob]:
    if not tables:
        return jobs
    tables_lower = {t.lower() for t in tables}
    return [
        j for j in jobs
        if j.target_table.lower() not in tables_lower
        and j.oracle_table.lower() not in tables_lower
    ]


def parse_table_list(raw: str | None) -> set[str]:
    if not raw:
        return set()
    return {t.strip() for t in raw.split(",") if t.strip()}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="ETL Oracle TOTVS Protheus -> PostgreSQL."
    )
    parser.add_argument(
        "--lookback-days",
        type=int,
        default=7,
        help="Dias para filtro incremental. Use 0 para carga total sem filtro.",
    )
    parser.add_argument(
        "--tables",
        default="",
        help=(
            "Lista separada por virgulas de tabelas a executar "
            "(aceita target_table ou oracle_table). "
            "Exemplo: 'SC1,SC7,SE1' ou 'SC1010,SC7010'."
        ),
    )
    parser.add_argument(
        "--exclude-tables",
        default="",
        help="Lista separada por virgulas de tabelas a ignorar.",
    )
    parser.add_argument(
        "--export-xlsx",
        default="",
        help=(
            "Caminho do arquivo XLSX a gerar. Quando informado, extrai do Oracle "
            "sem conectar ou gravar no PostgreSQL."
        ),
    )
    parser.add_argument(
        "--delete-mode",
        action="store_true",
        help="Executa em modo de exclusao (apaga no Supabase as chaves extraidas do Oracle).",
    )
    return parser.parse_args()


def _run_job_with_retry(
    job: EtlJob,
    oracle_schema: str,
    lookback_days: int | None,
    job_index: int,
    total_jobs: int,
    is_delete_mode: bool = False,
) -> str | None:
    """Executa um job com retry (ate JOB_MAX_ATTEMPTS tentativas).

    Cada tentativa cria e descarta seu proprio engine PostgreSQL e
    conexao Oracle para garantir isolamento total.

    Retorna None em caso de sucesso, ou a mensagem de erro em caso de falha
    definitiva apos todas as tentativas.
    """
    last_exc: Exception | None = None

    for attempt in range(1, JOB_MAX_ATTEMPTS + 1):
        engine: Engine | None = None
        ora_conn: oracledb.Connection | None = None
        try:
            engine = create_engine(
                _sqlalchemy_database_url(require_postgres_database_url()),
                poolclass=QueuePool,
                pool_size=5,
                max_overflow=2,
                pool_pre_ping=True,
            )
            ora_conn = create_oracle_connection()
            run_job(engine, ora_conn, oracle_schema, job, lookback_days, is_delete_mode)
            logger.info(
                "Job %s/%s | oracle_table=%s | target=%s concluido com sucesso"
                " na tentativa %s/%s.",
                job_index, total_jobs,
                job.oracle_table, job.target_table,
                attempt, JOB_MAX_ATTEMPTS,
            )
            return None  # sucesso

        except Exception as exc:
            last_exc = exc
            if attempt < JOB_MAX_ATTEMPTS:
                wait = 10 * attempt
                logger.warning(
                    "Job oracle_table=%s target_table=%s falhou na tentativa"
                    " %s/%s. Aguardando %ss antes de tentar novamente. Erro: %s",
                    job.oracle_table, job.target_table,
                    attempt, JOB_MAX_ATTEMPTS, wait, exc,
                )
                time.sleep(wait)
            else:
                logger.exception(
                    "Job oracle_table=%s target_table=%s falhou em todas as"
                    " %s tentativas.",
                    job.oracle_table, job.target_table, JOB_MAX_ATTEMPTS,
                )

        finally:
            if ora_conn is not None:
                try:
                    ora_conn.close()
                except Exception:
                    pass
            if engine is not None:
                engine.dispose()

    return f"{type(last_exc).__name__}: {last_exc}"


def main() -> None:
    args = parse_args()
    lookback_days = None if args.lookback_days == 0 else args.lookback_days

    oracle_schema, jobs = load_jobs()

    include_tables = parse_table_list(args.tables)
    if include_tables:
        jobs = filter_jobs_by_tables(jobs, include_tables)
    else:
        jobs = exclude_jobs_by_tables(jobs, parse_table_list(args.exclude_tables))

    if not jobs:
        logger.info("Nenhum job selecionado; nada a executar.")
        return

    # A exportacao XLSX depende apenas do Oracle; o Supabase nao e consultado.
    preflight_oracle_connection()
    if args.export_xlsx:
        ora_conn = create_oracle_connection()
        try:
            export_jobs_to_excel(
                ora_conn,
                oracle_schema,
                jobs,
                lookback_days,
                Path(args.export_xlsx),
            )
        finally:
            ora_conn.close()
        return

    # Valida a conexao do destino antes de iniciar a carga no Supabase.
    _preflight_engine = create_engine(
        _sqlalchemy_database_url(require_postgres_database_url()),
        poolclass=QueuePool,
        pool_size=1,
        max_overflow=0,
        pool_pre_ping=True,
    )
    try:
        preflight_database_connection(_preflight_engine)
    finally:
        _preflight_engine.dispose()

    logger.info(
        "Jobs selecionados (%s no total): %s.",
        len(jobs),
        [f"{j.oracle_table}->{j.target_table}" for j in jobs],
    )

    failed_jobs: dict[str, str] = {}
    succeeded_jobs: list[str] = []

    try:
        for job_index, job in enumerate(jobs, start=1):
            logger.info("%-60s", "=" * 60)
            logger.info(
                "Iniciando job %s/%s | oracle_table=%s | target=%s.",
                job_index, len(jobs), job.oracle_table, job.target_table,
            )

            error_msg = _run_job_with_retry(
                job, oracle_schema, lookback_days, job_index, len(jobs), args.delete_mode
            )

            if error_msg is None:
                succeeded_jobs.append(job.target_table)
            else:
                failed_jobs[job.target_table] = error_msg

        logger.info("%-60s", "=" * 60)
        logger.info(
            "Pipeline concluido: %s jobs com sucesso, %s jobs com falha.",
            len(succeeded_jobs), len(failed_jobs),
        )
        if succeeded_jobs:
            logger.info("Sucesso: %s.", ", ".join(succeeded_jobs))
        if failed_jobs:
            logger.error("Falhas ao final do pipeline:")
            for tbl, err in failed_jobs.items():
                logger.error("  %s => %s", tbl, err)
            raise RuntimeError(
                f"Jobs com falha: {', '.join(failed_jobs)}"
            )

    except (KeyError, RuntimeError):
        logger.exception("Pipeline interrompido.")
        raise
    except Exception:
        logger.exception("Pipeline ETL interrompido.")
        raise


if __name__ == "__main__":
    main()
